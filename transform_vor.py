"""
transform_vor.py — Desktop GUI-инструмент трансформации иерархического ВОР в плоский список ЛЗК.

Запуск: python transform_vor.py
Зависимости: openpyxl (pip install openpyxl), tkinter (стандартная библиотека Python)

Бизнес-логика вынесена в vor_core.py и используется совместно с web-сервисом (app.py).
"""

import os
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from vor_core import (
        col_letter_to_index, is_hierarchy_num, detect_sheet,
        detect_data_start, detect_qty_col, detect_columns,
        transform,
    )
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Трансформация ВОР → Плоский список (ЛЗК)")
        self.resizable(False, False)
        self.configure(padx=16, pady=16)

        self._input_path = tk.StringVar()
        self._sheet_var = tk.StringVar()
        self._col_num = tk.StringVar(value="A")
        self._col_name = tk.StringVar(value="B")
        self._col_unit = tk.StringVar(value="C")
        self._col_qty = tk.StringVar(value="G")
        self._output_path = tk.StringVar()
        self._status = tk.StringVar(value="Выберите входной файл ВОР")
        self._sheets = []

        self._build_ui()
        self._check_openpyxl()

    # ------------------------------------------------------------------
    def _check_openpyxl(self):
        if not OPENPYXL_OK:
            messagebox.showerror(
                "Отсутствует зависимость",
                "Библиотека openpyxl не установлена.\n\n"
                "Установите её командой:\n  pip install openpyxl\n\n"
                "Затем перезапустите программу."
            )
            self.destroy()

    # ------------------------------------------------------------------
    def _build_ui(self):
        PAD = {"padx": 6, "pady": 4}

        # --- Входной файл ---
        frame_in = ttk.LabelFrame(self, text="Входной файл ВОР")
        frame_in.grid(row=0, column=0, sticky="ew", **PAD)
        ttk.Entry(frame_in, textvariable=self._input_path, width=58,
                  state="readonly").grid(row=0, column=0, padx=6, pady=6)
        ttk.Button(frame_in, text="Открыть…", command=self._browse_input).grid(
            row=0, column=1, padx=6, pady=6)

        # --- Лист ---
        frame_sheet = ttk.LabelFrame(self, text="Лист")
        frame_sheet.grid(row=1, column=0, sticky="ew", **PAD)
        self._sheet_cb = ttk.Combobox(frame_sheet, textvariable=self._sheet_var,
                                       width=40, state="readonly")
        self._sheet_cb.grid(row=0, column=0, padx=6, pady=6, sticky="w")
        ttk.Label(frame_sheet, text="(авто-определяется при открытии файла)",
                  foreground="gray").grid(row=0, column=1, padx=6)

        # --- Настройка колонок ---
        frame_cols = ttk.LabelFrame(self, text="Настройка колонок (буква A-Z или номер)")
        frame_cols.grid(row=2, column=0, sticky="ew", **PAD)

        col_defs = [
            ("№ п.п.:",     self._col_num,  0),
            ("Наименование:", self._col_name, 1),
            ("Ед. изм.:",    self._col_unit, 2),
            ("Кол-во итого:", self._col_qty,  3),
        ]
        for label, var, ci in col_defs:
            r, c = divmod(ci, 2)
            ttk.Label(frame_cols, text=label).grid(row=r, column=c * 2, padx=6, pady=4, sticky="e")
            ttk.Entry(frame_cols, textvariable=var, width=6).grid(
                row=r, column=c * 2 + 1, padx=6, pady=4, sticky="w")

        # --- Выходной файл ---
        frame_out = ttk.LabelFrame(self, text="Выходной файл")
        frame_out.grid(row=3, column=0, sticky="ew", **PAD)
        ttk.Entry(frame_out, textvariable=self._output_path, width=58).grid(
            row=0, column=0, padx=6, pady=6)
        ttk.Button(frame_out, text="Изменить…", command=self._browse_output).grid(
            row=0, column=1, padx=6, pady=6)

        # --- Прогресс + кнопка ---
        frame_act = ttk.Frame(self)
        frame_act.grid(row=4, column=0, sticky="ew", **PAD)
        self._progress = ttk.Progressbar(frame_act, length=400, mode="determinate")
        self._progress.grid(row=0, column=0, padx=6, pady=6)
        self._btn_run = ttk.Button(frame_act, text="Преобразовать",
                                    command=self._run, width=18)
        self._btn_run.grid(row=0, column=1, padx=6, pady=6)

        # --- Статус ---
        ttk.Label(self, textvariable=self._status,
                  relief="sunken", anchor="w", width=70).grid(
            row=5, column=0, sticky="ew", padx=6, pady=(4, 0))

        # --- Копирайт ---
        ttk.Label(
            self,
            text="Разработка: Руководитель отдела архитектуры и проектирования — Пахарев Кирилл  |  kpakharev@afid.ru",
            foreground="gray",
            font=("TkDefaultFont", 8),
            anchor="center",
        ).grid(row=6, column=0, sticky="ew", padx=6, pady=(6, 2))

    # ------------------------------------------------------------------
    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="Выберите файл ВОР",
            filetypes=[("Excel файлы", "*.xlsx *.xlsm"), ("Все файлы", "*.*")]
        )
        if not path:
            return
        self._input_path.set(path)
        self._status.set("Загрузка файла…")
        self.update_idletasks()
        self._load_file_info(path)

    def _load_file_info(self, path: str):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{e}")
            self._status.set("Ошибка открытия файла")
            return

        self._sheets = sheets
        self._sheet_cb["values"] = sheets

        # Авто-определение листа
        try:
            wb2 = openpyxl.load_workbook(path, data_only=True)
            best_sheet = detect_sheet(wb2)
            wb2.close()
        except Exception:
            best_sheet = None

        if best_sheet:
            self._sheet_var.set(best_sheet)
            self._status.set(f"Лист авто-определён: «{best_sheet}»")
            self._auto_detect_qty_col(path, best_sheet)
        elif sheets:
            self._sheet_var.set(sheets[0])
            self._status.set("Лист авто-определить не удалось — выберите вручную")
        else:
            self._status.set("В файле нет листов")

        # Выходной файл по умолчанию
        base = os.path.splitext(path)[0]
        self._output_path.set(base + "_ЛЗК.xlsx")

        # Обновление при смене листа вручную
        self._sheet_cb.bind("<<ComboboxSelected>>", self._on_sheet_changed)

    def _on_sheet_changed(self, _event=None):
        path = self._input_path.get()
        sheet = self._sheet_var.get()
        if path and sheet:
            self._auto_detect_qty_col(path, sheet)

    def _auto_detect_qty_col(self, path: str, sheet_name: str):
        """Попытаться авто-определить колонку Кол-во итого."""
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[sheet_name]
            col_num_idx = col_letter_to_index(self._col_num.get() or "A")
            idx = detect_qty_col(ws, col_num_idx)
            wb.close()
            if idx is not None:
                letter = get_column_letter(idx + 1)
                self._col_qty.set(letter)
        except Exception:
            pass

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            title="Сохранить как",
            defaultextension=".xlsx",
            filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")]
        )
        if path:
            self._output_path.set(path)

    # ------------------------------------------------------------------
    def _run(self):
        input_file = self._input_path.get()
        sheet_name = self._sheet_var.get()
        output_file = self._output_path.get()

        if not input_file:
            messagebox.showwarning("Нет файла", "Выберите входной файл ВОР.")
            return
        if not sheet_name:
            messagebox.showwarning("Нет листа", "Выберите лист для обработки.")
            return
        if not output_file:
            messagebox.showwarning("Нет пути", "Укажите путь для выходного файла.")
            return

        # Проверка колонок
        try:
            col_num  = col_letter_to_index(self._col_num.get())
            col_name = col_letter_to_index(self._col_name.get())
            col_unit = col_letter_to_index(self._col_unit.get())
            col_qty  = col_letter_to_index(self._col_qty.get())
        except Exception as e:
            messagebox.showerror("Ошибка колонок", f"Неверное обозначение колонки:\n{e}")
            return

        self._btn_run.config(state="disabled")
        self._progress["value"] = 0
        self._status.set("Обработка…")

        def worker():
            try:
                selected_cols = [col_name, col_unit, col_qty]
                count, mode, _ = transform(
                    input_file, sheet_name,
                    col_num, selected_cols, None,
                    output_file,
                    progress_callback=self._update_progress,
                )
                self.after(0, lambda c=count, md=mode, f=output_file: self._on_done(c, md, f))
            except Exception as e:
                msg = str(e)
                self.after(0, lambda m=msg: self._on_error(m))

        threading.Thread(target=worker, daemon=True).start()

    def _update_progress(self, current: int, total: int):
        pct = int(current / total * 100) if total else 0
        self.after(0, lambda v=pct: self._progress.__setitem__("value", v))

    def _on_done(self, count: int, mode: str, output_file: str):
        self._progress["value"] = 100
        self._btn_run.config(state="normal")
        self._status.set(f"Готово! Режим: {mode} | Записано строк: {count}")
        if messagebox.askyesno(
            "Готово",
            f"Трансформация завершена.\nРежим: {mode}\nЗаписано строк: {count}\n\nОткрыть выходной файл?"
        ):
            os.startfile(output_file)

    def _on_error(self, msg: str):
        self._progress["value"] = 0
        self._btn_run.config(state="normal")
        self._status.set("Ошибка!")
        messagebox.showerror("Ошибка при обработке", msg)


# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app = App()
    app.mainloop()
