"""
app.py — Flask web-сервис трансформации ВОР → ЛЗК.

Запуск: python app.py
Порт:   6511

Зависимости: flask>=3.0, openpyxl>=3.1
"""

import os
from io import BytesIO

from flask import Flask, request, jsonify, send_file, render_template
import openpyxl
from openpyxl.utils import get_column_letter

from vor_core import (
    col_letter_to_index, detect_sheet, detect_qty_col,
    detect_columns, detect_name_col, transform,
)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB


@app.route("/")
def index():
    resp = render_template("index.html")
    from flask import make_response
    r = make_response(resp)
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    r.headers["Pragma"] = "no-cache"
    return r


@app.route("/detect", methods=["POST"])
def detect():
    """
    Принимает Excel-файл (multipart), возвращает JSON:
    {
        "sheets": ["Лист1", ...],
        "detected_sheet": "Лист1" | null,
        "qty_cols": {"Лист1": "G", ...},
        "sheet_columns": {
            "Лист1": [
                {"index": 1, "letter": "B", "header": "Наименование"},
                ...
            ]
        }
    }
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Файл не передан"}), 400

    try:
        data = file.read()
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    except Exception as e:
        return jsonify({"error": f"Не удалось открыть файл: {e}"}), 422

    sheets = wb.sheetnames
    detected = detect_sheet(wb)

    col_num_idx = 0  # колонка иерархии — всегда A (индекс 0)
    qty_cols = {}
    sheet_columns = {}
    name_cols = {}  # индекс колонки с наименованием (будет исключена из экспорта в режиме "только работы")

    for name in sheets:
        try:
            ws = wb[name]
            idx = detect_qty_col(ws, col_num_idx)
            qty_cols[name] = get_column_letter(idx + 1) if idx is not None else None
            sheet_columns[name] = detect_columns(ws, col_num_idx)
            name_cols[name] = detect_name_col(ws, col_num_idx)
        except Exception:
            qty_cols[name] = None
            sheet_columns[name] = []
            name_cols[name] = None

    wb.close()
    return jsonify({
        "sheets": sheets,
        "detected_sheet": detected,
        "qty_cols": qty_cols,
        "sheet_columns": sheet_columns,
        "name_cols": name_cols,
    })


@app.route("/transform", methods=["POST"])
def do_transform():
    """
    Принимает multipart/form-data:
      file      — Excel-файл ВОР
      sheet     — имя листа
      col_num   — буква колонки иерархии (A по умолчанию)
      cols      — comma-separated буквы выбранных колонок (напр. "B,C,G")

    Возвращает Excel-файл результата (attachment).
    Ничего не пишет на диск.
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Файл не передан"}), 400

    cols_raw = request.form.get("cols", "").strip()
    if not cols_raw:
        return jsonify({"error": "Не выбраны столбцы для экспорта"}), 400

    try:
        col_num = col_letter_to_index(request.form.get("col_num", "A"))
    except Exception as e:
        return jsonify({"error": f"Неверное обозначение колонки иерархии: {e}"}), 400

    try:
        selected_cols = [
            col_letter_to_index(c.strip())
            for c in cols_raw.split(",")
            if c.strip()
        ]
    except Exception as e:
        return jsonify({"error": f"Неверное обозначение колонки: {e}"}), 400

    try:
        data = file.read()
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)

        # Определяем лист по индексу (избегаем проблем с кодировкой спецсимволов в имени)
        sheet_idx_raw = request.form.get("sheet_idx", "")
        if sheet_idx_raw.isdigit():
            idx = int(sheet_idx_raw)
            sheet_name = wb.sheetnames[idx] if 0 <= idx < len(wb.sheetnames) else None
        else:
            sheet_name = request.form.get("sheet", "").strip() or None

        if not sheet_name:
            return jsonify({"error": "Не указан лист"}), 400

        ws = wb[sheet_name]
        # Получить заголовки для выбранных колонок
        all_cols = detect_columns(ws, col_num)
        col_header_map = {c["index"]: c["header"] for c in all_cols}
        col_headers = [
            col_header_map.get(ci, get_column_letter(ci + 1))
            for ci in selected_cols
        ]
        wb.close()

        count, mode, buf = transform(
            BytesIO(data), sheet_name,
            col_num, selected_cols, col_headers,
            output_target=None,
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 422

    original_name = file.filename or "vor"
    base = os.path.splitext(original_name)[0]
    out_name = f"{base}_ЛЗК.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=out_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=6511, debug=False)
