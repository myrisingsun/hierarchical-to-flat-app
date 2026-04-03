"""
vor_core.py — Ядро трансформации ВОР → ЛЗК (без GUI, без дисковых зависимостей).

Используется как transform_vor.py (desktop GUI), так и app.py (Flask web).
"""

import re
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Вспомогательные функции
# ---------------------------------------------------------------------------

HIERARCHY_RE = re.compile(r'^\d+(\.\d+)*$')


def col_letter_to_index(letter: str) -> int:
    """Буква колонки (A-Z) → 0-based индекс. Принимает также числовые строки."""
    letter = letter.strip().upper()
    if letter.isdigit():
        return int(letter) - 1
    if len(letter) == 1 and letter.isalpha():
        return ord(letter) - ord('A')
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result - 1


def is_hierarchy_num(value) -> bool:
    """Проверить, является ли значение номером в иерархии ('1', '1.1', '1.1.1')."""
    if value is None:
        return False
    return bool(HIERARCHY_RE.match(str(value).strip()))


def detect_sheet(wb) -> str | None:
    """Найти лист с иерархическими данными ВОР. Возвращает имя листа или None."""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        count = 0
        for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row), values_only=True):
            if row and is_hierarchy_num(row[0]):
                count += 1
                if count >= 3:
                    return sheet_name
    return None


def detect_data_start(ws, col_num: int) -> int:
    """Найти первую строку данных: строка, где col_num = номер уровня 1 ('1', '2', ...)."""
    first_level_re = re.compile(r'^\d+$')
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        val = row[col_num] if col_num < len(row) else None
        if val is not None and first_level_re.match(str(val).strip()):
            return i
    return 1


def detect_qty_col(ws, col_num: int) -> int | None:
    """
    Найти колонку с итоговым количеством материала/работы.
    Ищем в заголовочных строках ключевые слова, а также смотрим на строки данных.
    """
    data_start = detect_data_start(ws, col_num)
    priority_keywords = ('итого', 'расход', 'списан', 'списание')
    fallback_keywords = ('кол-во', 'количество')
    best_priority = None
    best_fallback = None
    for row in ws.iter_rows(min_row=max(1, data_start - 5), max_row=data_start - 1, values_only=True):
        for ci, cell in enumerate(row):
            if not cell:
                continue
            cell_low = str(cell).lower()
            if best_priority is None and any(k in cell_low for k in priority_keywords):
                best_priority = ci
            if best_fallback is None and any(k in cell_low for k in fallback_keywords):
                best_fallback = ci
    if best_priority is not None:
        return best_priority
    if best_fallback is not None:
        return best_fallback
    # Попытка 2: первая строка материала (col_num=None)
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        num = row[col_num] if col_num < len(row) else None
        if num is None:
            for ci in range(col_num + 3, min(len(row), col_num + 10)):
                if row[ci] is not None and isinstance(row[ci], (int, float)):
                    return ci
            break
    # Попытка 3: первая числовая ячейка в строке работы
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        num = row[col_num] if col_num < len(row) else None
        if is_hierarchy_num(num) and str(num).count('.') >= 2:
            for ci in range(col_num + 2, min(len(row), col_num + 8)):
                if row[ci] is not None and isinstance(row[ci], (int, float)):
                    return ci
            break
    return None


def detect_name_col(ws, col_num_idx: int) -> int | None:
    """
    Найти колонку с наименованием работы/материала.
    Это первая непустая колонка после col_num_idx в строке с иерархическим номером.
    """
    data_start = detect_data_start(ws, col_num_idx)
    for row in ws.iter_rows(min_row=data_start, max_row=min(data_start + 50, ws.max_row), values_only=True):
        num = row[col_num_idx] if col_num_idx < len(row) else None
        if is_hierarchy_num(num) and str(num).count('.') >= 1:
            for ci in range(col_num_idx + 1, min(len(row), col_num_idx + 6)):
                if row[ci] is not None and str(row[ci]).strip():
                    return ci
    return col_num_idx + 1  # fallback: следующая колонка


def detect_columns(ws, col_num_idx: int) -> list:
    """
    Вернуть список всех значимых колонок листа, кроме col_num_idx.

    Заголовки берутся из строки перед data_start. Если строки нет —
    используются буквы колонок.

    Возвращает: [{"index": int, "letter": str, "header": str}, ...]
    """
    data_start = detect_data_start(ws, col_num_idx)
    header_row_idx = data_start - 1

    # Считать строку заголовков
    col_headers = {}
    if header_row_idx >= 1:
        for row in ws.iter_rows(
            min_row=header_row_idx, max_row=header_row_idx, values_only=True
        ):
            for ci, cell in enumerate(row):
                if cell is not None:
                    text = str(cell).strip()
                    if text and text.lower() != 'none':
                        col_headers[ci] = text
            break

    # Определить максимальный используемый столбец из данных
    max_col = ws.max_column or 0

    result = []
    for i in range(max_col):
        if i == col_num_idx:
            continue
        letter = get_column_letter(i + 1)
        header = col_headers.get(i, letter)
        result.append({"index": i, "letter": letter, "header": header})

    # Убрать хвостовые колонки без заголовка и без данных
    # Оставляем все, у которых есть заголовок или есть данные в первых 50 строках
    data_cols = set()
    for row in ws.iter_rows(
        min_row=data_start, max_row=min(data_start + 50, ws.max_row), values_only=True
    ):
        for ci, cell in enumerate(row):
            if cell is not None and ci != col_num_idx:
                data_cols.add(ci)

    result = [
        col for col in result
        if col["index"] in col_headers or col["index"] in data_cols
    ]

    return result


def sheet_has_material_rows(ws, col_num: int, data_start: int) -> bool:
    """
    Проверить, есть ли на листе строки материалов.
    Материальная строка — та, где col_num пуст, но есть содержимое в других ячейках.
    """
    checked = 0
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        num = row[col_num] if col_num < len(row) else None
        if num is None:
            other_content = any(
                row[i] for i in range(len(row)) if i != col_num
            )
            if other_content:
                return True
        checked += 1
        if checked > 200:
            break
    return False


# ---------------------------------------------------------------------------
# Трансформация
# ---------------------------------------------------------------------------

def transform(
    input_source,
    sheet_name: str,
    col_num: int,
    selected_cols: list,
    col_headers: list | None = None,
    output_target=None,
    progress_callback=None,
) -> tuple:
    """
    Трансформирует иерархический ВОР в плоский список.

    input_source   — путь к файлу (str) или файлоподобный объект (BytesIO)
    sheet_name     — имя листа
    col_num        — 0-based индекс колонки с иерархическим номером
    selected_cols  — список 0-based индексов колонок для экспорта
    col_headers    — список заголовков для selected_cols (None → буквы колонок)
    output_target  — путь к файлу (str) или None (вернуть BytesIO)

    Возвращает (количество_строк, режим, buf_или_None).
    """
    wb = openpyxl.load_workbook(input_source, data_only=True)
    ws = wb[sheet_name]

    data_start = detect_data_start(ws, col_num)
    has_materials = sheet_has_material_rows(ws, col_num, data_start)
    mode = "с материалами" if has_materials else "только работы"

    # Подготовить заголовки
    if col_headers is None:
        col_headers = [get_column_letter(ci + 1) for ci in selected_cols]

    # В режиме "только работы" имя работы уже идёт в фиксированный столбец
    # "Наименование работ" — исключаем name_col из selected_cols чтобы не дублировать.
    # В режиме "с материалами" name_col в строках материалов содержит имя материала
    # (другие данные) — не исключаем.
    if not has_materials:
        name_col_idx = detect_name_col(ws, col_num)
        filtered = [
            (ci, hdr) for ci, hdr in zip(selected_cols, col_headers)
            if ci != name_col_idx
        ]
        if filtered:
            selected_cols, col_headers = zip(*filtered)
            selected_cols = list(selected_cols)
            col_headers = list(col_headers)
        else:
            selected_cols, col_headers = [], []

    current_section = ""
    current_work_num = ""
    current_work_name = ""
    rows_out = []

    total_rows = ws.max_row
    processed = 0

    def get_vals(row):
        return [
            row[ci] if ci < len(row) else None
            for ci in selected_cols
        ]

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        processed += 1
        if progress_callback and processed % 100 == 0:
            progress_callback(processed, total_rows)

        if not row:
            continue

        num = row[col_num] if col_num < len(row) else None

        if num is None:
            if not has_materials:
                continue
            # Материальная строка — есть хоть какое-то содержимое
            other_content = any(row[i] for i in range(len(row)) if i != col_num)
            if other_content:
                vals = get_vals(row)
                rows_out.append(
                    (current_work_num, current_section, current_work_name) + tuple(vals)
                )
        elif is_hierarchy_num(num):
            num_str = str(num).strip()
            depth = num_str.count('.')
            # Имя берём из первого непустого значения после col_num
            name = None
            for ci in range(col_num + 1, min(len(row), col_num + 5)):
                if row[ci] is not None:
                    name = str(row[ci]).strip()
                    break
            if depth == 0:
                current_section = name or ""
            if depth >= 2:
                current_work_num = num_str
                current_work_name = name or ""
                if not has_materials:
                    vals = get_vals(row)
                    rows_out.append(
                        (current_work_num, current_section, current_work_name) + tuple(vals)
                    )
        else:
            if rows_out:
                break

    buf = _write_output(rows_out, col_headers, output_target)
    return len(rows_out), mode, buf


def _write_output(rows: list, col_headers: list, output_target) -> BytesIO | None:
    """
    Записать плоский список в Excel.
    output_target=None  → вернуть BytesIO (без диска)
    output_target=str   → сохранить в файл, вернуть None
    """
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "ЛЗК"

    fixed_headers = [
        "№ работ",
        "Раздел сметы",
        "Наименование работ",
    ]
    headers = fixed_headers + list(col_headers)

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_out.append(headers)
    for ci, _ in enumerate(headers, start=1):
        cell = ws_out.cell(row=1, column=ci)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws_out.row_dimensions[1].height = 30

    cell_align_wrap = Alignment(vertical="top", wrap_text=True)
    cell_align_center = Alignment(horizontal="center", vertical="top")

    for row in rows:
        ws_out.append(list(row))
        ri = ws_out.max_row
        ws_out.cell(ri, 1).alignment = cell_align_center   # № работ
        ws_out.cell(ri, 2).alignment = cell_align_wrap     # Раздел
        ws_out.cell(ri, 3).alignment = cell_align_wrap     # Наименование работ
        for ci in range(4, len(headers) + 1):
            ws_out.cell(ri, ci).alignment = cell_align_wrap

    # Ширины: фиксированные 3 + 20 для динамических
    fixed_widths = [10, 25, 50]
    for ci, w in enumerate(fixed_widths, start=1):
        ws_out.column_dimensions[get_column_letter(ci)].width = w
    for ci in range(len(fixed_widths) + 1, len(headers) + 1):
        ws_out.column_dimensions[get_column_letter(ci)].width = 20

    ws_out.freeze_panes = "A2"

    if output_target is None:
        buf = BytesIO()
        wb_out.save(buf)
        buf.seek(0)
        return buf
    else:
        wb_out.save(output_target)
        return None
